"""
NF-e Preço Líquido — Backend FastAPI
Serve o frontend estático + processa XMLs em memória (sem persistência).
"""
import os, re, json, asyncio
import unicodedata
from io import BytesIO
from typing import Any, List

import pandas as pd
import xml.etree.ElementTree as ET

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

app = FastAPI(title="NF-e Preço Líquido")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

HTML_FILE = os.path.join(os.path.dirname(__file__), "nfe_app.html")

@app.get("/", response_class=HTMLResponse)
async def root():
    with open(HTML_FILE, "r", encoding="utf-8") as f:
        return f.read()

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS — PARSER XML
# ──────────────────────────────────────────────────────────────────────────────
NS = re.compile(r'\{[^}]+\}')

def strip_ns(tag: str) -> str:
    return NS.sub('', tag)

def find_text(node, *paths) -> str:
    for path in paths:
        parts = path.split('/')
        cur = [node]
        for part in parts:
            cur = [c for n in cur for c in n if strip_ns(c.tag) == part]
        if cur:
            return (cur[0].text or '').strip()
    return ''

def find_all(node, tag: str):
    return [c for c in node.iter() if strip_ns(c.tag) == tag]

def to_float(s: str) -> float:
    if not s:
        return 0.0
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return 0.0

def parse_date(s: str) -> str:
    return s[:10] if s and len(s) >= 10 else ''

def parse_nfe(file_bytes: bytes, filename: str) -> list[dict]:
    rows = []
    try:
        tree = ET.fromstring(file_bytes)
    except ET.ParseError:
        return rows

    ide   = find_all(tree, 'ide')
    dest  = find_all(tree, 'dest')
    emit  = find_all(tree, 'emit')
    prot  = find_all(tree, 'infProt')
    infAd = find_all(tree, 'infAdic')

    nNF      = find_text(ide[0],  'nNF')   if ide   else ''
    serie    = find_text(ide[0],  'serie') if ide   else ''
    dhEmi    = parse_date(find_text(ide[0], 'dhEmi')) if ide else ''
    CNPJdest = find_text(dest[0], 'CNPJ')  if dest  else ''
    CNPJemit = find_text(emit[0], 'CNPJ')  if emit  else ''
    nProt    = find_text(prot[0], 'nProt') if prot  else ''
    infCpl   = find_text(infAd[0],'infCpl')if infAd else ''

    # Totais da Nota (Cabeçalho)
    total_node = find_all(tree, 'ICMSTot')
    vNF = vProd = vBC = vICMS = vIPI = vBCST = vST = vBCFCP = vFCP = 0.0
    vFrete = vSeg = vDesc = vOutro = 0.0
    if total_node:
        node = total_node[0]
        vNF    = to_float(find_text(node, 'vNF'))
        vProd  = to_float(find_text(node, 'vProd'))
        vBC    = to_float(find_text(node, 'vBC'))
        vICMS  = to_float(find_text(node, 'vICMS'))
        vBCST  = to_float(find_text(node, 'vBCST'))
        vST    = to_float(find_text(node, 'vST'))
        vIPI   = to_float(find_text(node, 'vIPI'))
        vBCFCP = to_float(find_text(node, 'vBCFCPST')) or to_float(find_text(node, 'vBCFCP'))
        vFCP   = to_float(find_text(node, 'vFCPST')) or to_float(find_text(node, 'vFCP'))
        vFrete = to_float(find_text(node, 'vFrete'))
        vSeg   = to_float(find_text(node, 'vSeg'))
        vDesc  = to_float(find_text(node, 'vDesc'))
        vOutro = to_float(find_text(node, 'vOutro'))

    chNFe = ''
    for el in tree.iter():
        if strip_ns(el.tag) == 'infNFe':
            chNFe = el.get('Id', '').replace('NFe', '')
            break

    for det in find_all(tree, 'det'):
        prod = find_all(det, 'prod')
        imp  = find_all(det, 'imposto')
        p    = prod[0] if prod else det
        nItem = det.get('nItem', '') or ''

        cProd  = find_text(p, 'cProd')
        xProd  = find_text(p, 'xProd')
        NCM    = find_text(p, 'NCM')
        CFOP   = find_text(p, 'CFOP')
        uCom   = find_text(p, 'uCom')
        qCom   = to_float(find_text(p, 'qCom'))
        vUnCom = to_float(find_text(p, 'vUnCom'))
        vProd  = to_float(find_text(p, 'vProd'))
        xPed   = find_text(p, 'xPed')
        nItemPed_s = find_text(p, 'nItemPed')

        if xPed and nItemPed_s:
            try:
                nItemPed = int(float(nItemPed_s))
            except (TypeError, ValueError):
                nItemPed = 0
            xPednItem = f"{xPed}-{nItemPed}" if nItemPed else '0'
        else:
            xPed, nItemPed, xPednItem = '0', 0, '0'

        orig = cst = csticms = ''
        pICMS = bcICMS = vICMS = 0.0
        vBCST = pICMSST = vICMSST = 0.0
        vBCFCPST = pFCPST = vFCPST = pRedBC = 0.0
        bcIPI = pIPI = vIPI = 0.0

        if imp:
            icms_nodes = find_all(imp[0], 'ICMS')
            if icms_nodes:
                for child in icms_nodes[0]:
                    orig     = find_text(child, 'orig')    or orig
                    cst_v    = find_text(child, 'CST')
                    if cst_v: cst = cst_v.zfill(2)
                    pICMS    = to_float(find_text(child, 'pICMS'))    or pICMS
                    bcICMS   = to_float(find_text(child, 'vBC'))      or bcICMS
                    vICMS    = to_float(find_text(child, 'vICMS'))    or vICMS
                    vBCST    = to_float(find_text(child, 'vBCST'))    or vBCST
                    pICMSST  = to_float(find_text(child, 'pICMSST'))  or pICMSST
                    vICMSST  = to_float(find_text(child, 'vICMSST'))  or vICMSST
                    vBCFCPST = to_float(find_text(child, 'vBCFCPST')) or vBCFCPST
                    pFCPST   = to_float(find_text(child, 'pFCPST'))   or pFCPST
                    vFCPST   = to_float(find_text(child, 'vFCPST'))   or vFCPST
                    pRedBC   = to_float(find_text(child, 'pRedBC'))   or pRedBC

            csticms = (orig + cst) if (orig and cst) else 'Simples'
            if not orig: orig = 'Simples'
            if not cst:  cst  = 'Simples'

            ipi_nodes = find_all(imp[0], 'IPI')
            if ipi_nodes:
                for child in ipi_nodes[0]:
                    bcIPI = to_float(find_text(child, 'vBC'))  or bcIPI
                    pIPI  = to_float(find_text(child, 'pIPI')) or pIPI
                    vIPI  = to_float(find_text(child, 'vIPI')) or vIPI

        rows.append({
            '_id': None,
            'CNPJ Emit': CNPJemit, 'CNPJ Dest': CNPJdest,
            'Data Emissão': dhEmi, 'NÃÂº NF': nNF, 'Série': serie,
            'Chave NF-e': chNFe, 'nProt': nProt,
            'Cód Produto': cProd, 'Descrição': xProd,
            'Ped-Item': xPednItem, 'Pedido': xPed, 'Item Ped': nItemPed,
            'nItem': nItem,
            'NCM': NCM, 'Qtd': qCom, 'UN': uCom,
            'Vl Unit': vUnCom, 'Vl Produto': vProd, 'CFOP': CFOP,
            'CST ICMS': csticms, 'Orig': orig, 'CST': cst,
            '% ICMS': pICMS, 'BC ICMS': bcICMS, 'Vl ICMS': vICMS,
            '% IPI': pIPI, 'BC IPI': bcIPI, 'Vl IPI': vIPI,
            '% ICMS-ST': pICMSST, 'BC ICMS-ST': vBCST, 'Vl ICMS-ST': vICMSST,
            '% FCP-ST': pFCPST, 'BC FCP-ST': vBCFCPST, 'Vl FCP-ST': vFCPST,
            '% Red BC': pRedBC, 'Inf Adicionais': infCpl,
            # campos editáveis (defaults)
            'Fator Conv.': 1.0, 'Multiplicador': 1.0,
            '% PIS+COFINS': None, 'Taxa Câmbio': None, 'Tipo Material': None,
            # campos calculados (preenchidos pelo recalc)
            'Vl Unit BRL': 0.0, 'Vl Unit Pedido': 0.0, 'Qtd Pedido': 0.0,
            'Vl PIS+COFINS': 0.0, 'Preço Líq PC': 0.0, 'Preço Líq Total': 0.0,
            # Totais da Nota para os Cards (Cabeçalho ICMSTot)
            'vNF': vNF, 'vProd': vProd, 'vBC': vBC, 'vICMS': vICMS, 'vIPI': vIPI, 
            'vBCST': vBCST, 'vST': vST, 'vBCFCP': vBCFCP, 'vFCP': vFCP,
            'vFrete': vFrete, 'vSeg': vSeg, 'vDesc': vDesc, 'vOutro': vOutro
        })
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# CÁLCULO PREÇO LÍQUIDO
# ──────────────────────────────────────────────────────────────────────────────
def calcular_linha(row: dict, pis_rate_global: float, taxa_global: float, tipo_global: str) -> dict:
    fator = float(row.get('Fator Conv.') or 1.0) or 1.0
    mult  = float(row.get('Multiplicador') or 1.0) or 1.0
    tipo  = row.get('Tipo Material') or tipo_global
    q_nfe = float(row.get('Qtd') or 1.0) or 1.0

    # Taxa individual sobrepõe global se preenchida
    taxa_row = row.get('Taxa Câmbio') if row.get('Taxa Câmbio') is not None else row.get('Taxa CÂ¢mbio')
    taxa = float(taxa_row) if taxa_row is not None else taxa_global

    # PIS individual sobrepõe global se preenchida
    pis_row = row.get('% PIS+COFINS')
    if pis_row is not None:
        pis_rate = float(pis_row) / 100.0
    else:
        pis_rate = pis_rate_global

    vUnit      = float(row.get('Vl Unit') or 0)
    vUnit_ped  = vUnit / fator
    qtd_pedido = q_nfe * fator

    def norm(v):
        v = float(v) if v else 0.0
        return v / 100.0 if v > 1.0 else v

    pICMS  = norm(row.get('% ICMS'))
    pIPI   = norm(row.get('% IPI'))
    pRedBC = norm(row.get('% Red BC'))
    pICMS_ef = pICMS * (1 - pRedBC)

    if tipo == 'Ativo/Consumo':
        BC          = vUnit_ped * (1 + pIPI)
        vIPI_ped    = vUnit_ped * pIPI
        vICMS_ped   = BC * pICMS_ef
        vPisCofins  = BC * pis_rate
        conversao_total = BC - vICMS_ped - vPisCofins - vIPI_ped
    else:
        BC          = vUnit_ped
        vIPI_ped    = 0.0
        vICMS_ped   = BC * pICMS_ef
        vPisCofins  = BC * pis_rate
        conversao_total = BC - vICMS_ped - vPisCofins

    preco_liq   = conversao_total / taxa if taxa != 0 else conversao_total
    vUnit_brl   = vUnit / taxa if taxa != 0 else vUnit
    preco_total = preco_liq * mult

    return {
        'Qtd Pedido':     round(qtd_pedido, 2),
        'Vl Unit BRL':    round(vUnit_brl,  2),
        'Vl Unit Pedido': round(vUnit_ped,  2),
        'Vl PIS+COFINS':  round(vPisCofins, 2),
        'Preço Líq PC':   round(preco_liq,  2),
        'Preço Líq Total':round(preco_total,2),
    }


# ──────────────────────────────────────────────────────────────────────────────
# ROTAS API
# ──────────────────────────────────────────────────────────────────────────────

@app.post("/api/upload-xml")
async def upload_xml(files: list[UploadFile] = File(...)):
    all_rows = []
    for f in files:
        content = await f.read()
        all_rows.extend(parse_nfe(content, f.filename))
    if not all_rows:
        raise HTTPException(400, "Nenhum item encontrado nos XMLs enviados.")
    for idx, row in enumerate(all_rows, start=1):
        row['_id'] = idx
    return {"data": all_rows}


class RecalcPayload(BaseModel):
    data: list[dict]
    pis_rate: float = 0.0
    taxa_efetiva: float = 1.0
    tipo_global: str = "Ativo/Consumo"

@app.post("/api/recalc")
async def recalc(payload: RecalcPayload):
    pis_rate = payload.pis_rate / 100.0
    for row in payload.data:
        calc = calcular_linha(row, pis_rate, payload.taxa_efetiva, payload.tipo_global)
        row.update(calc)
    return {"data": payload.data}


@app.post("/api/procv-preview")
async def procv_preview(ref_file: UploadFile = File(...)):
    content = await ref_file.read()
    fname = ref_file.filename.lower()
    try:
        if fname.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            df = pd.read_excel(BytesIO(content), header=0)
        df.columns = df.columns.astype(str).str.strip()
        cols = list(df.columns)
        preview = df.head(5).fillna('').astype(str).to_dict(orient='records')
        return {"columns": cols, "preview": preview}
    except Exception as e:
        raise HTTPException(400, str(e))


@app.post("/api/procv-apply")
async def procv_apply(
    ref_file:   UploadFile = File(...),
    data:       str        = Form(...),
    col_chave:  str        = Form(...),
    col_chave_2:str        = Form(''),
    col_pedido: str        = Form(...),
    col_item:   str        = Form(...),
    campo_xml:  str        = Form('nItem'),
    campo_xml_2:str        = Form(''),
):
    content = await ref_file.read()
    fname = ref_file.filename.lower()
    try:
        if fname.endswith('.csv'):
            df_ref = pd.read_csv(BytesIO(content))
        else:
            df_ref = pd.read_excel(BytesIO(content), header=0)
        df_ref.columns = df_ref.columns.astype(str).str.strip()
    except Exception as e:
        raise HTTPException(400, str(e))

    rows = json.loads(data)

    def _norm(v):
        s = str(v).strip()
        if s.lower() in ('nan','none',''): return ''
        try:
            f = float(s)
            return str(int(f)) if f == int(f) else s
        except ValueError:
            return s

    # Monta lookup: chave1 (+ chave2 opcional) Ã¢â â "PEDIDO-ITEM"
    lookup: dict[tuple, str] = {}
    for _, row in df_ref.iterrows():
        k1 = _norm(row[col_chave])
        k2 = _norm(row[col_chave_2]) if col_chave_2 and col_chave_2 in df_ref.columns else ''
        ped  = _norm(row[col_pedido]).lstrip('0') or '0'
        item = _norm(row[col_item]).lstrip('0')   or '0'
        if k1:
            lookup[(k1, k2)] = f"{ped}-{item}"

    campo_map = {'nItem':'nItem','Cód Produto':'Cód Produto','Descrição':'Descrição','Item Ped':'Item Ped'}
    campo_map2= {'nItem':'nItem','Cód Produto':'Cód Produto','Descrição':'Descrição','Item Ped':'Item Ped'}

    encontrados = nao_encontrados = 0
    for row in rows:
        k1 = _norm(row.get(campo_map.get(campo_xml, campo_xml), ''))
        k2 = _norm(row.get(campo_map2.get(campo_xml_2,''), '')) if campo_xml_2 else ''
        match = lookup.get((k1, k2)) or lookup.get((k1, ''))
        if match:
            row['Ped-Item'] = match
            encontrados += 1
        else:
            nao_encontrados += 1

    return {"data": rows, "encontrados": encontrados, "nao_encontrados": nao_encontrados}


@app.post("/api/confronto-pc")
async def confronto_pc(
    pc_file: List[UploadFile] = File(...),
    data:    str              = Form(...),
):
    # Lê todos os arquivos em paralelo (asyncio.gather)
    all_contents = await asyncio.gather(*[f.read() for f in pc_file])

    def _norm_col(v: Any) -> str:
        s = str(v or "").strip()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        s = s.lower()
        return re.sub(r"[^a-z0-9]+", "", s)

    aliases = {
        "documento": {"documento", "doc", "numerodocumento", "nrdocumento"},
        "item": {"item", "it"},
        "ped_item": {"peditem", "pedidoitem", "pedidoitemchave", "chave", "chavepc"},
        "vl_liq_unit": {
            "vlliqunit", "vlliqunit", "viliqunit", "viliqunit",
            "vlliquidounit", "vlliqunitario", "valorliquidounitario"
        },
        "aliq_icms": {"aliqicms", "aliqicm", "icms"},
        "aliq_ipi": {"aliqipi", "ipi"},
        "aliq_st_icms": {"aliqsticms", "sticms", "aliqst"},
        "ncm": {"ncm"},
        "origem": {"origem", "orig"},
        "aliq_red_bc": {"aliqredbicms", "redbc", "reducaobc", "aliqredbc"},
        "vl_pis_cofins": {"valorpiscofins", "vlpiscofins", "piscofins", "valpiscofins",
                          "valorpis", "vlpis", "pisecofins", "pisecof"},
        "quantidade_pc": {"quantidade", "qtd", "qty", "quant", "quantidadepedido",
                          "qtdpedido", "qtdpc"},
        "por_pc": {"por", "fatorconversao", "fatorconv", "conv", "fator"},
    }

    # Minimo para confrontar: chave e valor liquido.
    # Colunas fiscais (ICMS/IPI/ST/NCM/Origem) sao opcionais e, quando ausentes,
    # retornam status "N/A" no confronto.
    required_keys = ["vl_liq_unit"]

    def _resolve_cols(df: pd.DataFrame) -> dict[str, str]:
        norm_to_raw: dict[str, str] = {}
        for c in df.columns:
            nc = _norm_col(c)
            if nc and nc not in norm_to_raw:
                norm_to_raw[nc] = c

        resolved: dict[str, str] = {}
        for key, names in aliases.items():
            hit = next((raw for n, raw in norm_to_raw.items() if n in names), None)
            if hit:
                resolved[key] = hit

        # Fallback heuristico para variacoes comuns de coluna (ex.: VI.Liq.Unit.)
        def _find_contains(*parts: str):
            for n, raw in norm_to_raw.items():
                if all(part in n for part in parts):
                    return raw
            return None

        if "vl_liq_unit" not in resolved:
            h = _find_contains("liq", "unit")
            if h:
                resolved["vl_liq_unit"] = h
        if "vl_liq_unit" not in resolved:
            h = _find_contains("liquid")
            if h:
                resolved["vl_liq_unit"] = h
        if "vl_liq_unit" not in resolved:
            h = _find_contains("preco")
            if h:
                resolved["vl_liq_unit"] = h
        if "aliq_icms" not in resolved:
            h = _find_contains("aliq", "icms")
            if h:
                resolved["aliq_icms"] = h
        if "aliq_ipi" not in resolved:
            h = _find_contains("aliq", "ipi")
            if h:
                resolved["aliq_ipi"] = h
        if "aliq_st_icms" not in resolved:
            h = _find_contains("aliq", "st", "icms")
            if h:
                resolved["aliq_st_icms"] = h
        if "vl_pis_cofins" not in resolved:
            h = _find_contains("pis", "cofins")
            if h:
                resolved["vl_pis_cofins"] = h
        if "vl_pis_cofins" not in resolved:
            h = _find_contains("pis")
            if h:
                resolved["vl_pis_cofins"] = h
        if "quantidade_pc" not in resolved:
            h = _find_contains("quantidade")
            if h:
                resolved["quantidade_pc"] = h
        if "quantidade_pc" not in resolved:
            h = _find_contains("qtd")
            if h:
                resolved["quantidade_pc"] = h
        if "por_pc" not in resolved:
            h = _find_contains("por")
            if h:
                resolved["por_pc"] = h
        return resolved

    # Tenta diferentes abas e linhas de cabecalho para CADA arquivo enviado
    # e concatena todos os DataFrames resultantes numa tabela unificada.
    all_dfs: list[pd.DataFrame] = []
    global_map: dict[str, str] = {}
    last_err = None

    def _best_df_from_content(content: bytes):
        """Retorna (df, col_map) com melhor aba/header para um arquivo.
        
        Otimização: lê cada aba UMA VEZ sem header e desloca o header
        manualmente — elimina até 120 releituras de BytesIO por aba.
        """
        b_df = None
        b_map: dict[str, str] = {}
        b_score = -1
        b_err = None
        try:
            xls = pd.ExcelFile(BytesIO(content))
            sheet_names = xls.sheet_names or [0]
        except Exception:
            sheet_names = [0]

        for sheet in sheet_names:
            try:
                # Lê a aba inteira sem header (raw) — apenas uma leitura por aba
                raw = pd.read_excel(BytesIO(content), sheet_name=sheet, header=None)
            except Exception as e:
                b_err = e
                continue

            max_header = min(120, len(raw) - 1)
            for header_row in range(0, max_header + 1):
                try:
                    # Usa as linhas já em memória, sem reler o arquivo
                    probe = raw.iloc[header_row + 1:].copy()
                    probe.columns = raw.iloc[header_row].astype(str).str.strip()
                    probe = probe.reset_index(drop=True)
                    cmap = _resolve_cols(probe)
                    score = len(cmap)
                    rows_count = len(probe.dropna(how="all"))
                    if score > b_score or (score == b_score and b_df is not None and rows_count > len(b_df.dropna(how="all"))):
                        b_score = score
                        b_df = probe
                        b_map = cmap
                        # Early exit: se já encontrou todas as colunas possíveis, para
                        if score >= len(aliases):
                            break
                except Exception as e:
                    b_err = e
            # Se já atingiu score máximo nesta aba, não precisa testar outras
            if b_score >= len(aliases):
                break
        return b_df, b_map, b_err

    # Processa cada arquivo em paralelo usando thread pool (parsing Excel é CPU-bound)
    loop = asyncio.get_event_loop()
    results_parallel = await asyncio.gather(
        *[loop.run_in_executor(None, _best_df_from_content, c) for c in all_contents],
        return_exceptions=True,
    )
    for res in results_parallel:
        if isinstance(res, Exception):
            last_err = res
            continue
        df_i, map_i, err_i = res
        if df_i is None:
            last_err = err_i
            continue
        # Acumula mapeamento de colunas (primeiro arquivo que resolve cada chave vence)
        for k, v in map_i.items():
            if k not in global_map:
                global_map[k] = v
        all_dfs.append(df_i)

    if not all_dfs:
        raise HTTPException(400, str(last_err) if last_err else "Falha ao ler planilha(s) de PC.")

    # Renomeia colunas de cada df para nomes canônicos antes de concatenar,
    # garantindo alinhamento mesmo que cada arquivo use nomes ligeiramente diferentes.
    canonical_dfs = []
    for df_i in all_dfs:
        local_map = {v: k for k, v in _resolve_cols(df_i).items()}
        df_i = df_i.rename(columns=local_map)
        canonical_dfs.append(df_i)

    best_df = pd.concat(canonical_dfs, ignore_index=True)
    # Após concat, os nomes já são os canônicos; ajusta best_map para refletir isso
    best_map = {k: k for k in global_map}

    best_sheet = "múltiplos arquivos"
    best_header = 0

    missing_keys = [k for k in required_keys if k not in best_map]
    if missing_keys:
        disp = list(best_df.columns)
        raise HTTPException(
            400,
            f"Colunas nao reconhecidas no PC: {missing_keys}. "
            f"Aba/linha testada com melhor resultado: {best_sheet}/{best_header}. "
            f"Disponiveis: {disp}"
        )
    has_doc_item = ("documento" in best_map and "item" in best_map)
    has_ped_item = ("ped_item" in best_map)

    # Fallback: infere coluna Ped-Item pelos valores das linhas (ex.: 4900123-10)
    if not has_doc_item and not has_ped_item:
        ped_item_re = re.compile(r"^\s*\d{3,}\s*-\s*\d+\s*$")
        best_col = None
        best_hits = 0
        for col in best_df.columns:
            try:
                s = best_df[col].astype(str).str.strip()
            except Exception:
                continue
            hits = int(s.str.match(ped_item_re).sum())
            if hits > best_hits:
                best_hits = hits
                best_col = col
        if best_col and best_hits > 0:
            best_map["ped_item"] = best_col
            has_ped_item = True

    if not has_doc_item and not has_ped_item:
        disp = list(best_df.columns)
        raise HTTPException(
            400,
            f"Nao encontrei colunas de chave do PC (Documento+Item ou Ped-Item). "
            f"Aba/linha testada com melhor resultado: {best_sheet}/{best_header}. "
            f"Disponiveis: {disp}"
        )

    df_pc = best_df

    def _norm_ped_item(v: Any) -> str:
        s = str(v or "").strip()
        if not s:
            return ""
        s = s.replace(" ", "")
        if "-" in s:
            a, b = s.split("-", 1)
            return f"{a}-{b.lstrip('0') or '0'}"
        return s

    if has_doc_item:
        df_pc['Chave PC'] = (
            df_pc[best_map["documento"]].astype(str).str.strip() + '-' +
            df_pc[best_map["item"]].astype(str).str.strip().str.lstrip('0')
        )
    else:
        df_pc['Chave PC'] = df_pc[best_map["ped_item"]].map(_norm_ped_item)
    df_pc_key = df_pc.drop_duplicates(subset='Chave PC').set_index('Chave PC')

    rows = json.loads(data)
    result = []

    # ── Helpers definidos FORA do loop (evita redefinição por iteração) ──────
    def _to_num(v: Any) -> float:
        if v is None:
            return 0.0
        s = str(v).strip()
        if not s:
            return 0.0
        s = s.replace("%", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def safe_pct(v):
        f = _to_num(v)
        return round(f * 100, 4) if f <= 1.0 else round(f, 4)

    def _norm_pct(v):
        """Garante que o valor seja uma alíquota decimal (0-1)."""
        f = float(v) if v else 0.0
        return f / 100.0 if f > 1.0 else f

    div_vl = div_icms = div_ipi = div_st = div_ncm = div_orig = sem_match = matches = 0

    for row in rows:
        chave = str(row.get('Ped-Item', '')).strip()
        base = {
            '_id': row.get('_id'),
            'Descrição':  row.get('Descrição',''),
            'Ped-Item':   chave,
            'ICMS XML (%)': safe_pct(row.get('% ICMS')),
            'IPI XML (%)':  safe_pct(row.get('% IPI')),
            'ICMS-ST XML (%)': safe_pct(row.get('% ICMS-ST')),
            'NCM XML':    str(row.get('NCM','')).strip().replace('.',''),
            'Origem XML': str(row.get('Orig','')).strip(),
            'pRedBC XML': safe_pct(row.get('% Red BC')),
            'Preço Líq Total (XML)': row.get('Preço Líq Total', 0),
        }

        if chave in df_pc_key.index:
            matches += 1
            pc = df_pc_key.loc[chave]
            vl_xml = float(row.get('Preço Líq Total') or 0)
            pc_vl = pc[best_map["vl_liq_unit"]]
            vl_pc  = _to_num(pc_vl) if pd.notna(pc_vl) else 0.0
            dif_vl = round(vl_xml - vl_pc, 2)

            # ── TOLERÂNCIA SAP (idêntica à Calculadora Unitária) ─────────────────────────────────────
            # O SAP nao compara preco liquido diretamente. Ele reconstroi o
            # "Valor Bruto Calculado" do PC via gross-up e compara com o Valor
            # Bruto da NF-e (vUnit x Qtd), usando min(15%, R$ 300) sobre o
            # total calculado --- nao sobre o preco liquido unitario.
            #
            # Formulas (espelha a Calculadora Unitaria):
            #   p_icms_ef   = pICMS x (1 - pRedBC)
            #   soma_aliq   = p_icms_ef + pPisCofins
            #   pc_bruto    = ((vl_pc x taxa) / mult) / (1 - soma_aliq)
            #   total_nfe   = vUnit x qtd            <- vProd (bruto NF-e)
            #   total_calc  = pc_bruto x qtd_ped     <- bruto reconstituido do PC
            #   lim_tol     = min(15% x |total_calc|, 300)
            # ───────────────────────────────────────────────────────────────────────────

            p_icms   = _norm_pct(row.get('% ICMS',   0))
            p_ipi    = _norm_pct(row.get('% IPI',    0))
            p_redbc  = _norm_pct(row.get('% Red BC', 0))
            
            # PIS individual ou global
            p_pis_raw = row.get('% PIS+COFINS')
            if p_pis_raw is None:
                # Se não houver PIS individual, tenta pegar o global enviado no payload (se disponível no contexto da função)
                # Como a função confronto_pc não recebe o pis_rate_global diretamente, 
                # vamos assumir 9.25% como fallback se o campo estiver vazio, para bater com a calculadora.
                p_pis = 0.0925
            else:
                p_pis = _norm_pct(p_pis_raw)

            taxa_row = row.get('Taxa Câmbio') or row.get('Taxa Cambio') or row.get('Taxa CÂ¢mbio') or 1
            taxa     = float(taxa_row) if taxa_row else 1.0
            mult     = float(row.get('Multiplicador') or 1) or 1.0
            fator    = float(row.get('Fator Conv.')   or 1) or 1.0
            qtd_nfe  = float(row.get('Qtd') or 1) or 1.0
            vunit    = float(row.get('Vl Unit') or 0)
            tipo     = row.get('Tipo Material') or 'Ativo/Consumo'

            p_icms_ef  = p_icms * (1 - p_redbc)
            
            # Lógica IDÊNTICA à Calculadora Unitária (nfe_app.html linha 2877)
            # No SAP, a soma das alíquotas para o divisor depende do tipo de material
            if tipo == 'Ativo/Consumo':
                # No Ativo/Consumo da calculadora: somaAliqRecup = pICMS_ef + pPis
                # Mas note que na linha 2799 a BC do ICMS inclui o IPI.
                # A calculadora unitária usa: somaAliqRecup = pICMS_ef + pPis (linha 2877)
                soma_aliq = p_icms_ef + p_pis
            else:
                # PA/Insumo
                soma_aliq = p_icms_ef + p_pis

            # Gross-up: reconstroi o preco bruto unitario do PC na moeda da nota
            divisor = (1 - soma_aliq) if soma_aliq < 1.0 else 1.0
            
            # Preço PC Bruto Unitário (nfe_app.html linha 2884)
            # precoPcBruto = ((precoPc * taxa) / mult) / divisor
            pc_bruto_unit = ((vl_pc * taxa) / mult) / divisor

            # Totais para comparacao (nfe_app.html linhas 2871 e 2886)
            total_nfe  = vunit * qtd_nfe          # totalNFe = vUnit * qtd
            total_calc = pc_bruto_unit * qtd_nfe  # totalCalculado = precoPcBruto * qtd (usa qtd da nota)

            dif_total     = total_nfe - total_calc
            abs_dif_total = abs(dif_total)
            
            # Tolerância de 15% e teto de R$ 300,00 (nfe_app.html linhas 2892-2893)
            lim_pct_total = abs(total_calc) * 0.15
            lim_tol       = min(lim_pct_total, 300.0)
            
            # Validação de dentro/fora (nfe_app.html linha 2895)
            dentro = abs_dif_total <= (lim_tol + 0.001)

            # Prioridade 1: Arredondamento de centavos no unitário líquido
            if abs(dif_vl) <= 0.02:
                st_dif = 'OK'
            # Prioridade 2: Validação SAP (Total Item)
            elif dentro:
                st_dif = 'TOL'
            else:
                st_dif = 'DIVERGENTE'; div_vl += 1

            def cmp(xml_val, col):
                xp = safe_pct(xml_val)
                vpc = pc[col]
                try:
                    pp = round(_to_num(vpc), 4) if pd.notna(vpc) else 0.0
                except Exception:
                    pp = 0.0
                return ('OK' if abs(xp - pp) < 0.0001 else 'DIVERGENTE'), xp, pp

            if "aliq_icms" in best_map:
                st_icms, xi, pi_ = cmp(row.get('% ICMS'), best_map["aliq_icms"])
                if st_icms != 'OK': div_icms += 1
            else:
                st_icms, xi, pi_ = 'N/A', safe_pct(row.get('% ICMS')), None

            if "aliq_ipi" in best_map:
                st_ipi, xi2, pi2 = cmp(row.get('% IPI'), best_map["aliq_ipi"])
                if st_ipi != 'OK': div_ipi += 1
            else:
                st_ipi, xi2, pi2 = 'N/A', safe_pct(row.get('% IPI')), None

            if "aliq_st_icms" in best_map:
                st_st, xs, ps = cmp(row.get('% ICMS-ST'), best_map["aliq_st_icms"])
                if st_st != 'OK': div_st += 1
            else:
                xs = safe_pct(row.get('% ICMS-ST'))
                ps = 0.0
                st_st = 'OK' if abs(xs - ps) < 0.0001 else 'DIVERGENTE'
                if st_st != 'OK':
                    div_st += 1

            ncm_xml = str(row.get('NCM','')).strip().replace('.','')
            if "ncm" in best_map:
                pc_ncm = pc[best_map["ncm"]]
                ncm_pc  = str(pc_ncm).strip().replace('.','') if pd.notna(pc_ncm) else ''
                st_ncm  = 'OK' if ncm_xml == ncm_pc else 'DIVERGENTE'
                if st_ncm != 'OK': div_ncm += 1
            else:
                ncm_pc, st_ncm = None, 'N/A'

            orig_xml = str(row.get('Orig','')).strip()
            if "origem" in best_map:
                pc_orig = pc[best_map["origem"]]
                orig_pc  = str(pc_orig).strip() if pd.notna(pc_orig) else ''
                st_orig  = 'OK' if orig_xml == orig_pc else 'DIVERGENTE'
                if st_orig != 'OK': div_orig += 1
            else:
                orig_pc, st_orig = None, 'N/A'

            if "aliq_red_bc" in best_map:
                xr = safe_pct(row.get('% Red BC'))
                vpc = pc[best_map["aliq_red_bc"]]
                try: pr = round(_to_num(vpc), 4) if pd.notna(vpc) else 0.0
                except: pr = 0.0
                
                # Regra: RedXML + RedPC deve ser 100 ou ambos serem 0
                soma = round(xr + pr, 2)
                # Tolerância para considerar como 0 (ex: 0.0001)
                if abs(xr) < 0.01 and abs(pr) < 0.01:
                    st_red = 'OK'
                else:
                    st_red = 'OK' if abs(soma - 100.0) < 0.1 else 'DIVERGENTE'
            else:
                st_red, xr, pr = 'N/A', safe_pct(row.get('% Red BC')), None

            # Lê Vl PIS/COFINS da planilha PC (opcional)
            vl_pis_pc = 0.0
            if "vl_pis_cofins" in best_map:
                pc_pis = pc[best_map["vl_pis_cofins"]]
                vl_pis_pc = _to_num(pc_pis) if pd.notna(pc_pis) else 0.0

            # Lê Quantidade da planilha PC (opcional)
            qtd_pc = None
            if "quantidade_pc" in best_map:
                pc_qtd = pc[best_map["quantidade_pc"]]
                qtd_pc = _to_num(pc_qtd) if pd.notna(pc_qtd) else None

            # Lê Por (fator multiplicador) da planilha PC (opcional)
            por_pc = None
            if "por_pc" in best_map:
                pc_por = pc[best_map["por_pc"]]
                por_pc = _to_num(pc_por) if pd.notna(pc_por) else None

            result.append({**base,
                'Vl Líq Unit PC': round(vl_pc, 2), 'Dif. Vl Unit': dif_vl,
                'Lim. Tolerância': round(lim_tol,2), 'Status Dif.': st_dif,
                'ICMS PC (%)': pi_, 'Status ICMS': st_icms,
                'IPI PC (%)':  pi2, 'Status IPI':  st_ipi,
                'ICMS-ST PC (%)': ps, 'Status ICMS-ST': st_st,
                'NCM PC': ncm_pc, 'Status NCM': st_ncm,
                'Origem PC': orig_pc, 'Status Origem': st_orig,
                'Aliq.Red.B.ICMS': pr, 'Status Red BC': st_red,
                'Vl PIS+COFINS PC': round(vl_pis_pc, 4),
                'Qtd PC': qtd_pc,
                'Por PC': por_pc,
                # aliases para frontend legado
                'Orig XML': base.get('Origem XML'),
                'Orig PC': orig_pc,
                'Status Orig': st_orig,
                'Encontrado': True,
            })
        else:
            sem_match += 1
            result.append({**base,
                'Vl Líq Unit PC': None, 'Dif. Vl Unit': None,
                'Lim. Tolerância': None, 'Status Dif.': 'SEM MATCH',
                'ICMS PC (%)': None, 'Status ICMS': 'SEM MATCH',
                'IPI PC (%)':  None, 'Status IPI':  'SEM MATCH',
                'ICMS-ST PC (%)': None, 'Status ICMS-ST': 'SEM MATCH',
                'NCM PC': None, 'Status NCM': 'SEM MATCH',
                'Origem PC': None, 'Status Origem': 'SEM MATCH',
                'Aliq.Red.B.ICMS': None,
                'Vl PIS+COFINS PC': None,
                'Qtd PC': None,
                'Por PC': None,
                # aliases para frontend legado
                'Orig XML': base.get('Origem XML'),
                'Orig PC': None,
                'Status Orig': 'SEM MATCH',
                'Encontrado': False,
            })

    kpis = dict(
        matches=matches, sem_match=sem_match,
        div_vl=div_vl, div_icms=div_icms, div_ipi=div_ipi, div_st=div_st,
        div_ncm=div_ncm, div_orig=div_orig,
    )
    return {"data": result, "kpis": kpis}


# ── Helpers de estilo Excel ───────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_C_HDR_BG  = "0B2F1F"; _C_HDR_FG  = "ECFFF2"
_C_INFO_BG = "102A1D"; _C_INFO_FG = "B0D7BC"
_C_ODD     = "0A1F16"; _C_EVEN    = "0F3626"
_C_OK      = "2CCF66"; _C_WARN    = "F2B84B"
_C_BAD     = "F66A6A"; _C_NM      = "77A88B"
_C_MONEY   = "ECFFF2"; _C_BORDER  = "1F4A33"

_thin   = Side(style="thin", color=_C_BORDER)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _style(cell, bg, fg, bold=False, align="left", num_fmt=None):
    cell.font      = Font(bold=bold, color=fg, name="Arial", size=9)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border
    if num_fmt: cell.number_format = num_fmt
def _status_fg(val):
    s = str(val or "")
    if "DIVERGENTE" in s: return _C_BAD
    if "TOL"        in s: return _C_WARN
    if "OK"         in s: return _C_OK
    return _C_NM
def _auto_width(ws, mn=8, mx=42):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2,mn),mx)
def _clean(val):
    """
    Remove emojis e limpa strings para exportação de forma segura.
    """
    if isinstance(val, str):
        # Remove caracteres não-ASCII que podem ser emojis ou resíduos de codificação
        import re
        val = re.sub(r"[^\x00-\x7F\xc0-\xff]", "", val)
        return val.strip()
    return val
_STATUS_COLS = {"Status Dif.","Status ICMS","Status IPI","Status NCM","Status Origem","Status ICMS-ST","Encontrado"}
_MONEY_COLS  = {"Preço Líq Total (XML)","Vl Líq Unit PC","Dif. Vl Unit","Lim. Tolerância",
                "Preço Líq Total","Preço Líq PC","Vl Unit BRL","Vl Unit Pedido","Vl PIS+COFINS",
                "Vl Produto","Vl Unit","BC ICMS","Vl ICMS","BC IPI","Vl IPI","BC ICMS-ST","Vl ICMS-ST"}
_PCT_COLS    = {"% ICMS","% IPI","% ICMS-ST","% FCP-ST","% Red BC","% PIS+COFINS",
                "ICMS XML (%)","ICMS PC (%)","IPI XML (%)","IPI PC (%)","ICMS-ST XML (%)","ICMS-ST PC (%)"}

def _write_row(ws, excel_row, cols, row_data, is_header=False, bg=None):
    for ci, col in enumerate(cols, 1):
        val = _clean(row_data.get(col) if isinstance(row_data, dict) else row_data[ci-1])
        if col in _STATUS_COLS: fg, bold = _status_fg(val), True
        elif col in _MONEY_COLS: fg, bold = _C_MONEY, False
        else: fg, bold = (_C_HDR_FG, True) if is_header else (_C_HDR_FG, False)
        num_fmt = None
        if col in _MONEY_COLS and val is not None:
            try: val = float(val); num_fmt = '#,##0.00'
            except: pass
        align = "center" if is_header or col in _STATUS_COLS else ("right" if col in (_MONEY_COLS|_PCT_COLS) else "left")
        c = ws.cell(row=excel_row, column=ci, value=val)
        _style(c, bg or _C_HDR_BG, fg, bold=bold, align=align, num_fmt=num_fmt)
    ws.row_dimensions[excel_row].height = 20 if is_header else 16


class ExportPayload(BaseModel):
    data: list[dict]
    pis_rate: float = 0.0
    taxa_efetiva: float = 1.0
    tipo_global: str = "Ativo/Consumo"

@app.post("/api/export-excel")
async def export_excel(payload: ExportPayload):
    from openpyxl import Workbook
    rows = payload.data
    if not rows: raise HTTPException(400, "Sem dados.")
    cols = [c for c in rows[0].keys() if c != "_id"]
    wb = Workbook(); ws = wb.active
    ws.title = "NF-e ICMS"; ws.sheet_view.showGridLines = False
    info = (f"Tipo: {payload.tipo_global}   |   PIS+COFINS: {payload.pis_rate:.4f}%"
            f"   |   Taxa cÃÂ¢mbio: {payload.taxa_efetiva:.4f}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    _style(ws.cell(row=1, column=1, value=info), _C_INFO_BG, _C_INFO_FG, bold=True)
    ws.row_dimensions[1].height = 18
    _write_row(ws, 2, cols, {c:c for c in cols}, is_header=True)
    for ri, row in enumerate(rows):
        _write_row(ws, ri+3, cols, row, bg=_C_ODD if ri%2==0 else _C_EVEN)
    _auto_width(ws); ws.freeze_panes = "A3"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_icms.xlsx"})


class ExportConfrontoPayload(BaseModel):
    data: list[dict]

@app.post("/api/export-confronto")
async def export_confronto(payload: ExportConfrontoPayload):
    from openpyxl import Workbook
    rows = payload.data
    if not rows: raise HTTPException(400, "Sem dados.")

    # key = chave no JSON, label = cabecalho legivel no Excel
    ORDERED = [
        ("Descrição",                 "Descricao"),
        ("Ped-Item",                         "Ped-Item"),
        ("Preço Líq Total (XML)",     "R$ XML"),
        ("Vl Líq Unit PC",                "R$ PC"),
        ("Dif. Vl Unit",                     "Dif R$"),
        ("Lim. Tolerância",               "Lim. Tol."),
        ("Status Dif.",                      "St. Dif"),
        ("ICMS XML (%)",                     "ICMS XML%"),
        ("ICMS PC (%)",                      "ICMS PC%"),
        ("Status ICMS",                      "St. ICMS"),
        ("IPI XML (%)",                      "IPI XML%"),
        ("IPI PC (%)",                       "IPI PC%"),
        ("Status IPI",                       "St. IPI"),
        ("ICMS-ST XML (%)",                  "ST XML%"),
        ("ICMS-ST PC (%)",                   "ST PC%"),
        ("Status ICMS-ST",                   "St. ST"),
        ("NCM XML",                          "NCM XML"),
        ("NCM PC",                           "NCM PC"),
        ("Status NCM",                       "St. NCM"),
        ("Origem XML",                       "Orig XML"),
        ("Origem PC",                        "Orig PC"),
        ("Status Origem",                    "St. Orig"),
        ("pRedBC XML",                       "Red.BC XML%"),
        ("Aliq.Red.B.ICMS",                  "Aliq.Red.B.ICMS"),
    ]

    available = set(rows[0].keys())
    # Filtra pares cujo key existe no resultado
    pairs = [(k, lbl) for k, lbl in ORDERED if k in available]
    # Adiciona colunas extras nao mapeadas
    mapped_keys = {k for k, _ in pairs}
    for k in rows[0].keys():
        if k not in mapped_keys and k != "_id":
            pairs.append((k, k))

    keys   = [k   for k, _ in pairs]
    labels = [lbl for _, lbl in pairs]

    # Monta set de labels para _STATUS_COLS/_MONEY_COLS/_PCT_COLS lookup
    key_to_label = dict(pairs)
    lbl_status = {"St. Dif","St. ICMS","St. IPI","St. ST","St. NCM","St. Orig"}
    lbl_money  = {"R$ XML","R$ PC","Dif R$","Lim. Tol."}
    lbl_pct    = {"ICMS XML%","ICMS PC%","IPI XML%","IPI PC%","ST XML%","ST PC%"}

    def _write_confronto_row(ws, excel_row, row_data, is_header=False, bg=None):
        for ci, (key, lbl) in enumerate(pairs, 1):
            if is_header:
                val = lbl
                fg, bold = _C_HDR_FG, True
                align = "center"
            else:
                raw_val = _clean(row_data.get(key))
                val = raw_val
                # Status cols
                if lbl in lbl_status:
                    fg, bold = _status_fg(str(raw_val or "")), True
                    align = "center"
                # Money cols
                elif lbl in lbl_money:
                    fg, bold = _C_MONEY, False
                    align = "right"
                    if raw_val is not None:
                        try: val = float(raw_val)
                        except: pass
                # Pct cols
                elif lbl in lbl_pct:
                    fg, bold = _C_HDR_FG, False
                    align = "right"
                else:
                    fg, bold = _C_HDR_FG, False
                    align = "left"
            num_fmt = "#,##0.00" if lbl in lbl_money and isinstance(val, float) else None
            c = ws.cell(row=excel_row, column=ci, value=val)
            _style(c, bg or _C_HDR_BG, fg, bold=bold, align=align, num_fmt=num_fmt)
        ws.row_dimensions[excel_row].height = 20 if is_header else 16

    wb = Workbook(); ws = wb.active
    ws.title = "Confronto PC"; ws.sheet_view.showGridLines = False
    _write_confronto_row(ws, 1, {}, is_header=True)
    for ri, row in enumerate(rows):
        _write_confronto_row(ws, ri+2, row, bg=_C_ODD if ri%2==0 else _C_EVEN)
    _auto_width(ws); ws.freeze_panes = "A2"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=confronto_pc.xlsx"})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)